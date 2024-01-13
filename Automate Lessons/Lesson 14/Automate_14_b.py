import openpyxl,os

wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb['Sheet']
sheet['A1'].value = 42
sheet['A2'].value = 'Hello'

os.chdir(r'C:\Users\David Huang\Downloads\Python')

# openpyxl.load_workbook('example2.xlsx')
sheet2 = wb.create_sheet()
sheet2.title = 'My New Sheet Name'

wb.save('example2.xlsx')
wb.create_sheet(index=0, title='My Other Sheet')
wb.save('example3.xlsx')

