import openpyxl, os

os.chdir(r'C:\Users\David Huang\Downloads\Python')

wb = openpyxl.load_workbook('example.xlsx')
type(wb)
sheet = wb['Sheet1']
type(sheet)
wb.sheetnames
cell = sheet['A1']
print(cell.value) 
print(sheet.cell(row=1, column=2))

for i in range(1, sheet.max_row + 1):
    print(i, sheet.cell(row=i, column=2).value)